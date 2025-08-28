"""
LangChain agent to compute monthly VR using Gemini.

This version enables Pandas-based tools so the LLM can orchestrate DataFrame operations
for filtering, joining, aggregating, and computing fields. Excel I/O tools are still
available, and there is also a direct pandas_write_excel tool.

Environment:
- GEMINI_API_KEY must be set.

Inputs (expected in ./xlsx):
- BaseConsolidada.xlsx: consolidated base to read.
- VR MENSAL 05.2025.xlsx: template to match layout/columns.

Output:
- VR_MENSAL_05.2025_FINAL.xlsx in ./xlsx
"""
from __future__ import annotations

import os
import json
from datetime import datetime, date, time
import time as _time
from typing import List, Dict, Any, Optional
import logging
import unicodedata

try:
    # LangChain core (tools/messages). Optional for direct Excel tools usage.
    from langchain_core.tools import tool  # type: ignore
    from langchain_core.messages import HumanMessage, SystemMessage, AIMessage, ToolMessage  # type: ignore
    from langchain_core.runnables import RunnableConfig  # type: ignore
except Exception:  # pragma: no cover - allow import-free fallback
    # Minimal fallback stubs so this module can be imported without LangChain.
    def tool(name: str, return_direct: bool = False):
        def decorator(fn):
            class SimpleTool:
                def __init__(self, f):
                    self._f = f
                    self.name = name
                    self.return_direct = return_direct
                def invoke(self, args):
                    # Expect args as a dict; call underlying function by kwargs
                    try:
                        return self._f(**(args or {}))
                    except TypeError:
                        # Fallback: pass-through single positional if provided
                        if isinstance(args, dict) and len(args) == 1:
                            return self._f(next(iter(args.values())))
                        raise
                def __call__(self, *a, **kw):
                    return self._f(*a, **kw)
            return SimpleTool(fn)
        return decorator

    class HumanMessage:  # minimal stub
        def __init__(self, content: str):
            self.content = content

    class SystemMessage:
        def __init__(self, content: str):
            self.content = content

    class AIMessage:
        def __init__(self, content: str = "", tool_calls=None):
            self.content = content
            self.tool_calls = tool_calls or []

    class ToolMessage:
        def __init__(self, content: str, tool_call_id: str = ""):
            self.content = content
            self.tool_call_id = tool_call_id

    class RunnableConfig(dict):
        pass

try:
    # LangChain Google Gemini
    from langchain_google_genai import ChatGoogleGenerativeAI
except Exception as e:  # pragma: no cover - allow import-time check
    ChatGoogleGenerativeAI = None  # type: ignore

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill  # type: ignore
except Exception:
    raise RuntimeError("openpyxl is required for Excel I/O tools.")

# Optional: Pandas for DataFrame operations
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None  # type: ignore

# Optional: load .env if present
try:
    from dotenv import load_dotenv  # type: ignore
except Exception:
    load_dotenv = None  # will still support raw file fallback


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_DIR = os.path.join(BASE_DIR, "xlsx")
DEFAULT_TEMPLATE_PATH = os.path.join(XLSX_DIR, "VR MENSAL 05.2025.xlsx")
DEFAULT_INPUT_PATH = os.path.join(XLSX_DIR, "BaseConsolidada.xlsx")
DEFAULT_OUTPUT_PATH = os.path.join(XLSX_DIR, "VR_MENSAL_05.2025_FINAL.xlsx")
AUDIT_LOG = os.path.join(BASE_DIR, "run.log")

# Expected output columns and order
EXPECTED_OUTPUT_HEADERS: List[str] = [
    "Matricula",
    "Admissão",
    "Sindicato do Colaborador",
    "Competência",
    "Dias trabalhados",
    "Valor Diario VR",
    "TOTAL_Prestador",
    "Custo empresa",
    "Desconto profissional",
    "OBS GERAL",
]

# In-memory DataFrame registry for pandas_* tools
_DF_REGISTRY: Dict[str, Any] = {}
_DF_COUNTER: int = 0
_LAST_DF_ID: Optional[str] = None


def _new_df_id() -> str:
    global _DF_COUNTER
    _DF_COUNTER += 1
    return f"df{_DF_COUNTER}"


def _resolve_df(df_id: Optional[str]) -> Dict[str, Any]:
    """Return a dict with either {'df': DataFrame, 'df_id': id} or {'error': '...'}.
    Falls back to the last known df if df_id is missing/unknown.
    """
    if pd is None:
        return {"error": "pandas is not available in this environment."}
    global _LAST_DF_ID
    if df_id and df_id in _DF_REGISTRY:
        return {"df": _DF_REGISTRY[df_id], "df_id": df_id}
    # Fallback to last
    if _LAST_DF_ID and _LAST_DF_ID in _DF_REGISTRY:
        return {"df": _DF_REGISTRY[_LAST_DF_ID], "df_id": _LAST_DF_ID}
    # If there's exactly one df, use it
    if len(_DF_REGISTRY) == 1:
        only_id = next(iter(_DF_REGISTRY.keys()))
        _LAST_DF_ID = only_id
        return {"df": _DF_REGISTRY[only_id], "df_id": only_id}
    return {"error": f"Unknown df_id: {df_id if df_id else '<none>'}"}


def _resolve_path(path: str) -> str:
    if os.path.isabs(path):
        return path
    # allow relative to base dir
    return os.path.join(BASE_DIR, path)


def _sheet_to_rows(path: str, sheet_name: Optional[str] = None, max_rows: Optional[int] = None) -> List[List[Any]]:
    wb = load_workbook(filename=path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows: List[List[Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and i >= max_rows:
            break
        rows.append([cell for cell in row])
    return rows


def _json_safe(value: Any) -> Any:
    """Convert Excel cell values to JSON-safe types; dates/times -> ISO strings."""
    try:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat()
        # Leave numbers, bool, None as-is; convert other objects to strings
        if isinstance(value, (int, float, bool)) or value is None:
            return value
        # For formulas or unexpected types, stringify gracefully
        return str(value)
    except Exception:
        return str(value) if value is not None else None


def _detect_header_row(all_rows: List[List[Any]], max_scan: int = 10) -> int:
    """Return zero-based index of the header row by scanning the first rows for the most non-empty text cells."""
    if not all_rows:
        return 0
    best_idx = 0
    best_score = -1
    scan_limit = min(max_scan, len(all_rows))
    for i in range(scan_limit):
        row = all_rows[i]
        # score: number of non-empty cell strings
        score = 0
        for cell in row:
            s = str(cell).strip() if cell is not None else ""
            if s:
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _rows_to_workbook(headers: List[str], rows: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha1"
    # write header
    ws.append(headers)
    # write rows in header order
    for r in rows:
        ws.append([r.get(h, None) for h in headers])
    return wb


def _df_json_safe(v: Any) -> Any:
    """JSON-safe conversion for DataFrame cells."""
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    # Convert numpy types to native python
    try:
        import numpy as _np  # local to avoid hard dependency if missing
        if isinstance(v, (_np.integer,)):
            return int(v)
        if isinstance(v, (_np.floating,)):
            # Preserve two decimals for money-like floats when reasonable
            return float(v)
        if isinstance(v, (_np.bool_,)):
            return bool(v)
    except Exception:
        pass
    if pd is not None and isinstance(v, pd.Timestamp):  # type: ignore
        return v.isoformat()
    if pd is not None and pd.isna(v):  # type: ignore
        return None
    return v


def _df_to_table(df: Any, max_rows: Optional[int] = None, start_row: int = 0) -> Dict[str, Any]:
    headers = [str(c) for c in df.columns.tolist()]
    sub = df.iloc[start_row: start_row + max_rows if max_rows else None]
    rows = []
    for _, r in sub.iterrows():
        obj = {}
        for h, v in r.items():
            obj[str(h)] = _df_json_safe(v)
        rows.append(obj)
    return {"headers": headers, "rows": rows, "row_count": int(df.shape[0])}


def _norm(s: str) -> str:
    """Normalize a string: lowercase, remove accents, keep only alphanum."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower()
    return "".join(ch for ch in s if ch.isalnum())


def _find_col(df: Any, candidates: List[str]) -> Optional[str]:
    """Find the first existing column in df matching any of the candidate names (normalized)."""
    by_norm = { _norm(c): c for c in df.columns }
    for cand in candidates:
        n = _norm(cand)
        if n in by_norm:
            return by_norm[n]
    return None


# =============================
# Pandas Tools
# =============================

@tool("pandas_read_excel", return_direct=False)
def pandas_read_excel(
    file_path: str,
    sheet_name: Optional[str] = None,
    header: Optional[int] = 0,
) -> str:
    """Load an Excel sheet into a registered DataFrame. Returns JSON with {df_id, columns, shape}.
    - file_path: path to .xlsx
    - sheet_name: sheet name (optional)
    - header: header row index (0-based) or None to infer
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    path = _resolve_path(file_path)
    if not os.path.exists(path):
        return json.dumps({"error": f"File not found: {path}"})
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header, engine="openpyxl")
        # If multiple sheets returned, take first
        if isinstance(df, dict):
            # get active-like first
            first_key = list(df.keys())[0]
            df = df[first_key]
        df_id = _new_df_id()
        _DF_REGISTRY[df_id] = df
        global _LAST_DF_ID
        _LAST_DF_ID = df_id
        return json.dumps({
            "df_id": df_id,
            "columns": [str(c) for c in df.columns.tolist()],
            "shape": [int(df.shape[0]), int(df.shape[1])],
        })
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_preview", return_direct=False)
def pandas_preview(df_id: str, head: int = 5) -> str:
    """Return a preview of a registered DataFrame: columns, dtypes, head rows, shape."""
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        dtypes = {str(k): str(v) for k, v in df.dtypes.to_dict().items()}
        head_df = df.head(head)
        table = _df_to_table(head_df, max_rows=head, start_row=0)
        return json.dumps({
            "df_id": df_id,
            "columns": table["headers"],
            "dtypes": dtypes,
            "rows": table["rows"],
            "shape": [int(df.shape[0]), int(df.shape[1])],
        })
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_to_table", return_direct=False)
def pandas_to_table(df_id: str, max_rows: Optional[int] = None, start_row: int = 0) -> str:
    """Convert a registered DataFrame into a JSON table: {headers, rows, row_count}."""
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    try:
        table = _df_to_table(df, max_rows=max_rows, start_row=start_row)
        return json.dumps(table)
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_eval", return_direct=False)
def pandas_eval(df_id: str, expr: str, target_col: Optional[str] = None, inplace: bool = True) -> str:
    """Evaluate an expression using DataFrame.eval(). If target_col is provided, assigns the result to that column.
    Returns JSON with {df_id, columns} and optionally sample rows.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        if target_col:
            df[target_col] = df.eval(expr)
            out_id = df_id
            out_df = df
        else:
            series = df.eval(expr)
            out_df = df.copy()
            col = target_col or "_eval_result"
            out_df[col] = series
            if inplace:
                _DF_REGISTRY[df_id] = out_df
                out_id = df_id
            else:
                out_id = _new_df_id()
                _DF_REGISTRY[out_id] = out_df
        return json.dumps({
            "df_id": out_id,
            "columns": [str(c) for c in out_df.columns.tolist()],
        })
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_query", return_direct=False)
def pandas_query(df_id: str, query_expr: str, inplace: bool = True) -> str:
    """Filter rows using DataFrame.query(). Returns JSON with {df_id, shape}."""
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        filtered = df.query(query_expr)
        if inplace:
            _DF_REGISTRY[df_id] = filtered
            out_id = df_id
        else:
            out_id = _new_df_id()
            _DF_REGISTRY[out_id] = filtered
        return json.dumps({"df_id": out_id, "shape": [int(filtered.shape[0]), int(filtered.shape[1])]})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_merge", return_direct=False)
def pandas_merge(
    left_df_id: str,
    right_df_id: str,
    how: str = "left",
    on: Optional[str] = None,
    left_on: Optional[str] = None,
    right_on: Optional[str] = None,
    suffixes: Optional[str] = None,
    new_df_id: Optional[str] = None,
) -> str:
    """Merge two registered DataFrames. Returns JSON with {df_id, columns, shape}."""
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    lres = _resolve_df(left_df_id)
    rres = _resolve_df(right_df_id)
    if "error" in lres or "error" in rres:
        return json.dumps({"error": f"Unknown df_id for left or right"})
    ldf = lres["df"]
    rdf = rres["df"]
    try:
        suff = tuple(json.loads(suffixes)) if suffixes else ("_x", "_y")
        merged = ldf.merge(rdf, how=how, on=on, left_on=left_on, right_on=right_on, suffixes=suff)
        out_id = new_df_id or _new_df_id()
        _DF_REGISTRY[out_id] = merged
        return json.dumps({
            "df_id": out_id,
            "columns": [str(c) for c in merged.columns.tolist()],
            "shape": [int(merged.shape[0]), int(merged.shape[1])],
        })
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_groupby_agg", return_direct=False)
def pandas_groupby_agg(
    df_id: str,
    by_json: str,
    agg_spec_json: str,
    as_index: bool = False,
    new_df_id: Optional[str] = None,
) -> str:
    """Group by and aggregate. by_json: JSON array of column names. agg_spec_json: JSON mapping for aggregations.
    Example agg_spec: {"colA": "sum", "colB": ["mean", "max"]} or nested dict format.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    try:
        by = json.loads(by_json)
        agg_spec = json.loads(agg_spec_json)
        gb = df.groupby(by=by, as_index=as_index).agg(agg_spec)
        # Flatten columns if MultiIndex
        if isinstance(gb.columns, pd.MultiIndex):  # type: ignore
            gb.columns = ["_".join([str(c) for c in tup if str(c)]) for tup in gb.columns.tolist()]
        out_id = new_df_id or _new_df_id()
        _DF_REGISTRY[out_id] = gb.reset_index() if not as_index else gb
        out_df = _DF_REGISTRY[out_id]
        return json.dumps({
            "df_id": out_id,
            "columns": [str(c) for c in out_df.columns.tolist()],
            "shape": [int(out_df.shape[0]), int(out_df.shape[1])],
        })
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_fillna", return_direct=False)
def pandas_fillna(df_id: str, value: Optional[str] = None, columns_json: Optional[str] = None, method: Optional[str] = None, inplace: bool = True) -> str:
    """Fill NA values. value is used if provided; else method (e.g., 'ffill'/'bfill'). columns_json is an optional JSON array of columns to apply.
    Returns JSON with {df_id, shape}.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        target = df if inplace else df.copy()
        cols = json.loads(columns_json) if columns_json else None
        if cols is not None:
            target = target.copy()
            sub = target[cols]
            if value is not None:
                sub = sub.fillna(value)
            elif method is not None:
                sub = sub.fillna(method=method)
            target[cols] = sub
        else:
            if value is not None:
                target = target.fillna(value)
            elif method is not None:
                target = target.fillna(method=method)
        out_id = df_id if inplace else _new_df_id()
        _DF_REGISTRY[out_id] = target
        return json.dumps({"df_id": out_id, "shape": [int(target.shape[0]), int(target.shape[1])]})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_rename", return_direct=False)
def pandas_rename(df_id: str, columns_json: str, inplace: bool = True) -> str:
    """Rename columns using a JSON mapping {old: new}. Returns JSON with {df_id, columns}."""
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        mapping = json.loads(columns_json)
        target = df if inplace else df.copy()
        target = target.rename(columns=mapping)
        out_id = df_id if inplace else _new_df_id()
        _DF_REGISTRY[out_id] = target
        return json.dumps({"df_id": out_id, "columns": [str(c) for c in target.columns.tolist()]})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_write_excel", return_direct=True)
def pandas_write_excel(df_id: str, file_path: str, sheet_name: str = "Planilha1", index: bool = False) -> str:
    """Write a registered DataFrame to an Excel file and return the absolute path."""
    if pd is None:
        return f"ERROR: pandas is not available in this environment."
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return f"ERROR: {res['error']}"
    df = res["df"]
    # Remove lines with TOTAL_Prestador == 0 before writing
    try:
        if "TOTAL_Prestador" in df.columns:
            vals = pd.to_numeric(df["TOTAL_Prestador"], errors="coerce").fillna(0.0)
            df = df.loc[vals != 0.0].copy()
    except Exception:
        pass
    path = _resolve_path(file_path)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Compute totals and column positions for key columns
        totals_map: Dict[str, Optional[float]] = {}
        col_pos_map: Dict[str, int] = {}
        for col_name in ["TOTAL_Prestador", "Custo empresa", "Desconto profissional"]:
            if col_name in df.columns:
                try:
                    totals_map[col_name] = float(pd.to_numeric(df[col_name], errors="coerce").fillna(0).sum())
                except Exception:
                    totals_map[col_name] = None
                col_pos_map[col_name] = list(df.columns).index(col_name)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # Reserve row 1 for labels and row 2 for totals; start headers at row 3
            df.to_excel(writer, sheet_name=sheet_name, index=index, startrow=2)
            try:
                ws = writer.sheets[sheet_name]
                offset = 1 if index else 0
                # Totals (row 2)
                for name, sum_val in totals_map.items():
                    if sum_val is None:
                        continue
                    pos = col_pos_map.get(name)
                    if pos is None:
                        continue
                    col_excel = pos + 1 + offset
                    ws.cell(row=2, column=col_excel, value=sum_val)
                # Labels (row 1)
                label_map = {
                    "TOTAL_Prestador": "Total_Custo",
                    "Custo empresa": "Total_Custo_Empresa",
                    "Desconto profissional": "Total_Custo_Funcionario",
                }
                for col_name, label in label_map.items():
                    if col_name in df.columns:
                        pos = list(df.columns).index(col_name)
                        col_excel = pos + 1 + offset
                        cell = ws.cell(row=1, column=col_excel, value=label)
                        try:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7", end_color="FFFDE7")
                        except Exception:
                            pass
                # Currency format for totals row and data
                currency_cols = ["Valor Diario VR", "TOTAL_Prestador", "Custo empresa", "Desconto profissional"]
                fmt = '"R$" #,##0.00'
                header_row = 3
                first_data_row = header_row + 1  # 4
                last_row = ws.max_row
                for col_name in currency_cols:
                    if col_name in df.columns:
                        pos = list(df.columns).index(col_name)
                        col_excel = pos + 1 + (1 if index else 0)
                        ws.cell(row=2, column=col_excel).number_format = fmt
                        for r in range(first_data_row, last_row + 1):
                            ws.cell(row=r, column=col_excel).number_format = fmt
            except Exception:
                pass
        return path
    except Exception as e:
        # Fallback: if file is locked (Permission denied), write to a timestamped new file
        msg = str(e)
        if "Permission denied" in msg or "[Errno 13]" in msg:
            try:
                base, ext = os.path.splitext(path)
                alt = f"{base}.{int(_time.time())}{ext or '.xlsx'}"
                with pd.ExcelWriter(alt, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index, startrow=2)
                    try:
                        ws = writer.sheets[sheet_name]
                        offset = 1 if index else 0
                        # Totals row
                        for name in ["TOTAL_Prestador", "Custo empresa", "Desconto profissional"]:
                            if name in df.columns:
                                try:
                                    s = float(pd.to_numeric(df[name], errors="coerce").fillna(0).sum())
                                except Exception:
                                    s = None
                                if s is None:
                                    continue
                                pos = list(df.columns).index(name)
                                col_excel = pos + 1 + offset
                                ws.cell(row=2, column=col_excel, value=s)
                        # Labels row
                        label_map = {
                            "TOTAL_Prestador": "Total_Custo",
                            "Custo empresa": "Total_Custo_Empresa",
                            "Desconto profissional": "Total_Custo_Funcionario",
                        }
                        for col_name, label in label_map.items():
                            if col_name in df.columns:
                                pos = list(df.columns).index(col_name)
                                col_excel = pos + 1 + offset
                                cell = ws.cell(row=1, column=col_excel, value=label)
                                try:
                                    cell.font = Font(bold=True)
                                    cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7", end_color="FFFDE7")
                                except Exception:
                                    pass
                        # Currency formatting
                        currency_cols = ["Valor Diario VR", "TOTAL_Prestador", "Custo empresa", "Desconto profissional"]
                        fmt = '"R$" #,##0.00'
                        header_row = 3
                        first_data_row = header_row + 1
                        last_row = ws.max_row
                        for col_name in currency_cols:
                            if col_name in df.columns:
                                pos = list(df.columns).index(col_name)
                                col_excel = pos + 1 + (1 if index else 0)
                                ws.cell(row=2, column=col_excel).number_format = fmt
                                for r in range(first_data_row, last_row + 1):
                                    ws.cell(row=r, column=col_excel).number_format = fmt
                    except Exception:
                        pass
                return alt
            except Exception as e2:
                return f"ERROR: {e}; fallback failed: {e2}"
        return f"ERROR: {e}"


@tool("pandas_ensure_columns", return_direct=False)
def pandas_ensure_columns(df_id: str, headers_json: Optional[str] = None, fill_value: Any = None, drop_extras: bool = True) -> str:
    """Ensure the DataFrame has exactly the specified columns in the given order.
    - headers_json: JSON array of column names. If omitted, uses the built-in EXPECTED_OUTPUT_HEADERS.
    - fill_value: value to use when creating missing columns (default None).
    - drop_extras: if True, drops any columns not listed.
    Returns {df_id, columns} after enforcement.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        target_cols = json.loads(headers_json) if headers_json else EXPECTED_OUTPUT_HEADERS
        # Add missing
        for c in target_cols:
            if c not in df.columns:
                df[c] = fill_value
        # Drop extras
        if drop_extras:
            df = df[[c for c in target_cols if c in df.columns]]
        else:
            # Reorder; keep extras at end
            ordered = [c for c in target_cols if c in df.columns]
            extras = [c for c in df.columns if c not in target_cols]
            df = df[ordered + extras]
        _DF_REGISTRY[df_id] = df
        return json.dumps({"df_id": df_id, "columns": [str(c) for c in df.columns.tolist()]})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_auto_rename", return_direct=False)
def pandas_auto_rename(df_id: str) -> str:
    """Automatically rename common columns to match the required output names.
    Uses simple normalization and a synonym map to align to EXPECTED_OUTPUT_HEADERS.
    Returns {df_id, applied_mapping}.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        # Define synonyms -> target
        syn_map: Dict[str, str] = {}

        def add_synonyms(target: str, alts: List[str]):
            for a in alts:
                syn_map[_norm(a)] = target

        add_synonyms("Matricula", ["matricula", "MATRICULA"])    
        add_synonyms("Admissão", ["admissao", "admissão", "admissao data", "data admissao"])    
        add_synonyms("Sindicato do Colaborador", ["sindicato", "sindicato do colaborador"])    
        add_synonyms("Competência", ["competencia", "competência"])    
        add_synonyms("Dias trabalhados", ["dias", "dias trabalhados", "dias uteis", "dias_uteis"])    
        add_synonyms("Valor Diario VR", ["valor diario vr", "valor_diario_vr", "valor diário vr"])    
        add_synonyms("TOTAL_Prestador", ["total", "total prestador", "total_prestador"])    
        add_synonyms("Custo empresa", ["custo empresa", "custo_empresa"])    
        add_synonyms("Desconto profissional", ["desconto profissional", "desconto_profissional"])    
        add_synonyms("OBS GERAL", ["obs", "obs geral", "obs_geral"])    

        mapping: Dict[str, str] = {}
        for col in list(df.columns):
            n = _norm(str(col))
            target = syn_map.get(n)
            if target and str(col) != target:
                mapping[str(col)] = target
        if mapping:
            df = df.rename(columns=mapping)
            _DF_REGISTRY[df_id] = df
        return json.dumps({"df_id": df_id, "applied_mapping": mapping})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("pandas_populate_vr", return_direct=False)
def pandas_populate_vr(
    df_id: str,
    competencia: str = "05/2025",
    vigencia_start: str = "2025-04-15",
) -> str:
    """Populate required VR columns from common base fields.
    - competencia: e.g., '05/2025'.
    - vigencia_start: ISO date string 'YYYY-MM-DD' used for desligamento rule reference (15/05 in downstream logic).
    Fills/creates: [Matricula, Admissão, Sindicato do Colaborador, Competência, Dias trabalhados,
    Valor Diario VR, TOTAL_Prestador, Custo empresa, Desconto profissional, OBS GERAL].
    Uses columns when available: TOTAL_DIAS_UTEIS_SINDICATO, DIAS DE FÉRIAS, VALOR_DIARIO_VR, DATA DEMISSÃO, Sindicato, MATRÍCULA, Admissão.
    """
    if pd is None:
        return json.dumps({"error": "pandas is not available in this environment."})
    df = _DF_REGISTRY.get(df_id)
    res = _resolve_df(df_id)
    if "error" in res:
        return json.dumps(res)
    df = res["df"]
    df_id = res["df_id"]
    try:
        work = df.copy()
        # Ensure identity columns exist (try to map common sources)
        if "Matricula" not in work.columns:
            src = _find_col(work, ["Matricula", "MATRICULA", "matricula"]) or "Matricula"
            if src in work.columns and src != "Matricula":
                work["Matricula"] = work[src]
        if "Admissão" not in work.columns:
            src = _find_col(work, ["Admissão", "Admissao", "data admissao", "admissao"])
            if src:
                work["Admissão"] = work[src]
        if "Sindicato do Colaborador" not in work.columns:
            src = _find_col(work, ["Sindicato do Colaborador", "Sindicato"]) 
            if src:
                work["Sindicato do Colaborador"] = work[src]

        # Competência constant
        work["Competência"] = competencia

        # Source columns (robust by normalization lookup)
        col_diasuteis = _find_col(work, ["TOTAL_DIAS_UTEIS_SINDICATO", "dias uteis sindicato", "dias_uteis", "Dias trabalhados"]) 
        col_ferias = _find_col(work, ["DIAS DE FÉRIAS", "dias de ferias", "dias_ferias", "ferias"]) 
        col_valorvr = _find_col(work, ["Valor Diario VR", "VALOR_DIARIO_VR", "valor diario vr"]) 
        col_demissao = _find_col(work, ["DATA DEMISSÃO", "data demissao", "dt demissao", "demissao"]) 

        # Compute Dias trabalhados baseline
        dias = pd.to_numeric(work[col_diasuteis], errors="coerce") if col_diasuteis else 0
        if isinstance(dias, (int, float)):
            dias_series = pd.Series([dias] * len(work))
        else:
            dias_series = dias.fillna(0)
        if col_ferias:
            ferias = pd.to_numeric(work[col_ferias], errors="coerce").fillna(0)
            dias_series = (dias_series - ferias).clip(lower=0)

        # Desligamento rule: use cutoff = 15th of the competence month by default,
        # or honor a provided cutoff if available in the function arguments via caller.
        try:
            # Try to infer cutoff from competencia when not passed via config
            if competencia and (not isinstance(competencia, str) or len(competencia) < 7):
                competencia = str(competencia)
        except Exception:
            pass
        cutoff_str = None
        # Prefer an environment override via module-level last provided values in run_vr_agent metadata
        # Not directly accessible here, so rely on arguments; if no explicit cutoff, infer from competencia 'MM/YYYY'
        if competencia and isinstance(competencia, str) and "/" in competencia:
            try:
                mm, yyyy = competencia.split("/")
                cutoff_str = f"{int(yyyy):04d}-{int(mm):02d}-15"
            except Exception:
                cutoff_str = None
        # Fallback to May/2025 if nothing else
        if not cutoff_str:
            cutoff_str = "2025-05-15"
        cutoff = pd.to_datetime(cutoff_str, errors="coerce")
        obs = pd.Series([None] * len(work))
        if col_demissao:
            ddem = pd.to_datetime(work[col_demissao], errors="coerce")
            mask_zero = ddem.notna() & (ddem <= cutoff)
            dias_series = dias_series.mask(mask_zero, 0)
            obs = obs.mask(mask_zero, "Sem compra por desligamento <= 15/05")

        work["Dias trabalhados"] = dias_series.astype(float)

        # Valor Diario VR
        if col_valorvr:
            vr = pd.to_numeric(work[col_valorvr], errors="coerce").fillna(0.0)
        else:
            vr = pd.Series([0.0] * len(work))
        work["Valor Diario VR"] = vr.astype(float)

        # Totals
        total = (work["Dias trabalhados"].astype(float) * work["Valor Diario VR"].astype(float)).fillna(0.0)
        work["TOTAL_Prestador"] = total
        work["Custo empresa"] = (total * 0.80).round(2)
        work["Desconto profissional"] = (total * 0.20).round(2)
        work["OBS GERAL"] = obs

        _DF_REGISTRY[df_id] = work
        # Basic non-null stats to confirm population
        stats = {
            "rows": int(work.shape[0]),
            "non_null_TOTAL_Prestador": int(work["TOTAL_Prestador"].notna().sum()),
            "sum_TOTAL_Prestador": float(work["TOTAL_Prestador"].sum()),
        }
        return json.dumps({"df_id": df_id, "columns": [str(c) for c in work.columns.tolist()], "stats": stats})
    except Exception as e:
        return json.dumps({"error": str(e)})


@tool("read_excel_table", return_direct=False)
def read_excel_table(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_rows: Optional[int] = None,
    start_row: int = 1,
) -> str:
    """Read an Excel sheet and return JSON with {headers: [...], rows: [ {col: val} ... ]} (1-based rows).
    Parameters:
    - file_path: path to xlsx file (relative to project root or absolute).
    - sheet_name: optional sheet name; defaults to the active sheet.
    - max_rows: optional limit of total rows to load (including the header row when start_row=1).
    - start_row: 1-based index of the first data row to include in the response (header is always row 1).
    Notes:
    - The first row in the sheet is assumed to be the header and is always returned as 'headers'.
    - Data 'rows' start at 'start_row' (>=2 recommended to skip header), and up to 'max_rows'.
    """
    path = _resolve_path(file_path)
    if not os.path.exists(path):
        return json.dumps({"error": f"File not found: {path}"})
    all_rows = _sheet_to_rows(path, sheet_name=sheet_name, max_rows=max_rows)
    if not all_rows:
        return json.dumps({"error": "Empty sheet"})
    # Detect header row robustly
    hdr_idx = _detect_header_row(all_rows)
    headers = [str(h) if h is not None else "" for h in all_rows[hdr_idx]]
    data_rows = []
    # Default data starts after detected header; honor explicit start_row if it's beyond that
    default_sr = hdr_idx + 2  # 1-based Excel row for first data row
    sr = max(default_sr, int(start_row) if start_row else default_sr)
    # Convert to zero-based index
    start_idx = max(1, sr - 1)
    for row in all_rows[start_idx:]:
        obj = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            obj[h] = _json_safe(val)
        data_rows.append(obj)
    return json.dumps({"headers": headers, "rows": data_rows})


@tool("read_excel_headers", return_direct=False)
def read_excel_headers(file_path: str, sheet_name: Optional[str] = None) -> str:
    """Return only the header row (list of column names) from an Excel sheet as JSON {headers:[...]}."""
    path = _resolve_path(file_path)
    if not os.path.exists(path):
        return json.dumps({"error": f"File not found: {path}"})
    all_rows = _sheet_to_rows(path, sheet_name=sheet_name, max_rows=None)
    if not all_rows:
        return json.dumps({"error": "Empty sheet"})
    hdr_idx = _detect_header_row(all_rows)
    headers = [str(h) if h is not None else "" for h in all_rows[hdr_idx]]
    return json.dumps({"headers": headers})


@tool("write_excel_table", return_direct=True)
def write_excel_table(file_path: str, headers_json: str, rows_json: str) -> str:
    """Write an Excel file from headers and rows provided as JSON. Returns the absolute path.
    Parameters:
    - file_path: output path (.xlsx). Will be created or overwritten.
    - headers_json: JSON array of strings with the desired column order.
    - rows_json: JSON array of objects mapping column->value.
    This tool performs no calculations; it writes data produced by the LLM.
    """
    path = _resolve_path(file_path)
    try:
        headers: List[str] = json.loads(headers_json)
        rows: List[Dict[str, Any]] = json.loads(rows_json)
        # Remove lines with TOTAL_Prestador == 0 before writing
        try:
            if "TOTAL_Prestador" in headers:
                def _to_float(v):
                    try:
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            return 0.0
                        return float(v)
                    except Exception:
                        try:
                            s = str(v).replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
                            return float(s)
                        except Exception:
                            return 0.0
                rows = [r for r in rows if _to_float(r.get("TOTAL_Prestador")) != 0.0]
        except Exception:
            pass
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        # Reserve row 1 for labels and row 2 for totals; headers at row 3
        ws.append([None] * len(headers))  # row 1 labels
        ws.append([None] * len(headers))  # row 2 totals
        ws.append(headers)                # row 3 headers
        # Rows from row 4 onwards
        currency_cols = ["Valor Diario VR", "TOTAL_Prestador", "Custo empresa", "Desconto profissional"]
        for r in rows:
            out_row = []
            for h in headers:
                v = r.get(h, None)
                if h in currency_cols:
                    try:
                        v = float(v) if v is not None and str(v) != "" else None
                    except Exception:
                        v = None
                out_row.append(v)
            ws.append(out_row)
        # Totals in row 2 and labels in row 1
        try:
            key_cols = ["TOTAL_Prestador", "Custo empresa", "Desconto profissional"]
            for key in key_cols:
                if key in headers:
                    col_idx = headers.index(key) + 1
                    total_sum = 0.0
                    for r in rows:
                        v = r.get(key)
                        try:
                            total_sum += float(v) if v is not None and str(v) != "" else 0.0
                        except Exception:
                            pass
                    ws.cell(row=2, column=col_idx, value=total_sum)
            labels = {
                "TOTAL_Prestador": "Total_Custo",
                "Custo empresa": "Total_Custo_Empresa",
                "Desconto profissional": "Total_Custo_Funcionario",
            }
            for key, label in labels.items():
                if key in headers:
                    col_idx = headers.index(key) + 1
                    cell = ws.cell(row=1, column=col_idx, value=label)
                    try:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(fill_type="solid", start_color="FFFDE7", end_color="FFFDE7")
                    except Exception:
                        pass
        except Exception:
            pass
        # Apply BRL currency format
        try:
            fmt = '"R$" #,##0.00'
            header_row = 3
            first_data_row = header_row + 1
            last_row = ws.max_row
            for h in ["Valor Diario VR", "TOTAL_Prestador", "Custo empresa", "Desconto profissional"]:
                if h in headers:
                    col_idx = headers.index(h) + 1
                    ws.cell(row=2, column=col_idx).number_format = fmt
                    for r_idx in range(first_data_row, last_row + 1):
                        ws.cell(row=r_idx, column=col_idx).number_format = fmt
        except Exception:
            pass
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb.save(path)
        return path
    except Exception as e:
        return f"ERROR: {e}"


@tool("audit_log", return_direct=False)
def audit_log(step: str, details: str) -> str:
    """Append an audit entry with timestamp, step name, and details to run.log. Returns 'OK'."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {step}: {details}\n"
    try:
        with open(AUDIT_LOG, "a", encoding="utf-8") as f:
            f.write(line)
        return "OK"
    except Exception as e:
        return f"ERROR: {e}"


def _load_api_key_from_files() -> Optional[str]:
    """Attempt to load the Gemini API key from local files.
    Priority:
      1) .env file (GEMINI_API_KEY=...)
      2) gemini_api_key.txt (content is either the raw key or a 'GEMINI_API_KEY=...' line)
    Returns the key if found; otherwise None.
    """
    # 1) .env
    env_path = os.path.join(BASE_DIR, ".env")
    if os.path.exists(env_path):
        # Prefer python-dotenv if available
        if load_dotenv is not None:
            load_dotenv(env_path)
            v = os.getenv("GEMINI_API_KEY")
            if v:
                return v.strip()
        # Fallback: naive .env parsing
        try:
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if line.upper().startswith("GEMINI_API_KEY="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
        except Exception:
            pass
    # 2) gemini_api_key.txt
    txt_path = os.path.join(BASE_DIR, "gemini_api_key.txt")
    if os.path.exists(txt_path):
        try:
            with open(txt_path, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if not content:
                return None
            if content.upper().startswith("GEMINI_API_KEY="):
                return content.split("=", 1)[1].strip()
            return content
        except Exception:
            return None
    return None


def _load_gemini_api_key() -> Optional[str]:
    """Resolve Gemini API key preferring Streamlit secrets in Streamlit Cloud.
    Order:
      1) st.secrets["GEMINI_API_KEY"] (or "gemini_api_key") if Streamlit is available
      2) Environment variable GEMINI_API_KEY (or gemini_api_key)
      3) .env or gemini_api_key.txt via _load_api_key_from_files()
    """
    # 1) Streamlit secrets (if running inside Streamlit)
    try:
        import streamlit as st  # type: ignore
        v = None
        try:
            # st.secrets behaves like a dict; .get works
            v = st.secrets.get("GEMINI_API_KEY") or st.secrets.get("gemini_api_key")
        except Exception:
            # Older Streamlit versions may require indexing
            for k in ("GEMINI_API_KEY", "gemini_api_key"):
                try:
                    if k in st.secrets and st.secrets[k]:
                        v = st.secrets[k]
                        break
                except Exception:
                    pass
        if v:
            return str(v).strip()
    except Exception:
        pass

    # 2) Environment variables
    v = os.getenv("GEMINI_API_KEY") or os.getenv("gemini_api_key")
    if v:
        return v.strip()

    # 3) Local file fallbacks
    return _load_api_key_from_files()


def _get_logger() -> logging.Logger:
    """Get a configured logger that writes to run.log and console."""
    logger = logging.getLogger("vr_agent")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [AGENT] %(levelname)s: %(message)s")
    # File handler (append to the same run.log used by audit tool)
    fh = logging.FileHandler(AUDIT_LOG, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    # Console handler
    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    return logger


SYSTEM_PROMPT_BASE = """
You are a meticulous payroll operations agent tasked with calculating Vale Refeição (VR) for the month.

You are allowed and encouraged to use the provided pandas_* tools to perform data analysis, joins, filtering,
grouping, and calculations. Prefer DataFrame operations over manual reasoning for correctness and traceability.

Data Sources:
- Consolidated base: xlsx/BaseConsolidada.xlsx.
- Template: xlsx/VR MENSAL 05.2025.xlsx (use its header order and required fields).

Month Context:
{month_context}

Task Breakdown:
1) Load BaseConsolidada.xlsx via pandas_read_excel into a DataFrame. Use pandas_preview to inspect columns and types as needed.
2) Load the template header via read_excel_headers to determine exact output layout and column names to match in the final workbook.
3) Using pandas tools (eval/query/merge/groupby) or the helper pandas_populate_vr, compute for each collaborator, logging each step with audit_log:
   a. Contract status: determine if active. Start date = admission date; if missing, use 15/04 (YYYY-04-15).
   b. Working business days in the month: compute using provided spreadsheet fields only (e.g., per-sindicato business days and employee dates). Deduct proportionally for vacations and absences. Do not invent holidays; rely on the consolidated data (including per-sindicato days) and dates present.
   c. Desligamento rule: if the employee is desligado on or before 15/05 and status indicates "OK", exclude VR purchase (0); otherwise, consider full purchase for the period.
   d. Total VR value: number_of_business_days_worked * daily VR value (column with the daily VR amount). Keep monetary precision as string with two decimals.
4) Build a new table that exactly matches the template columns and naming. Include these computed fields:
   - valor_de_compra_vr (total to purchase)
   - custo_empresa (80% of total)
   - desconto_profissional (20% of total)
5) Enforce the exact column layout and naming before saving, using pandas_ensure_columns with these columns:
    [Matricula, Admissão, Sindicato do Colaborador, Competência, Dias trabalhados, Valor Diario VR, TOTAL_Prestador, Custo empresa, Desconto profissional, OBS GERAL].
6) Write the result to xlsx/VR_MENSAL_05.2025_FINAL.xlsx via either:
    - write_excel_table (using pandas_to_table to convert DataFrame to headers/rows), or
    - pandas_write_excel directly from a DataFrame.

Rules:
- If a required datum is missing, make a conservative assumption and log it using audit_log. Do not drop the row unless impossible to compute.
- Keep all calculations consistent and auditable; record a brief per-employee audit string summarizing the decision path.
- Output must be saved via write_excel_table or pandas_write_excel when producing the final workbook.
 - IMPORTANT: Do not finish with a normal chat response. You MUST conclude by calling either write_excel_table (with valid headers_json and rows_json) or pandas_write_excel so the file is written. Do not emit an empty final message.
"""


def build_agent(tools: Optional[List] = None):
    if ChatGoogleGenerativeAI is None:
        raise RuntimeError("langchain-google-genai is not installed. Add it to requirements and install.")
    # Resolve API key from Streamlit secrets first, then env/.env/txt
    api_key = _load_gemini_api_key()
    if not api_key:
        raise RuntimeError(
            "GEMINI_API_KEY ausente. Defina nos secrets do Streamlit Community Cloud (st.secrets), "
            "ou via variável de ambiente/.env; como último recurso, use gemini_api_key.txt no diretório do projeto."
        )

    llm = ChatGoogleGenerativeAI(
        model="gemini-1.5-pro",
        temperature=0.2,
        max_output_tokens=4096,
        convert_system_message_to_human=False,
        api_key=api_key,
    )

    bound = llm.bind_tools(tools or [
        # Excel I/O
        read_excel_table, read_excel_headers, write_excel_table, audit_log,
        # Pandas tools
        pandas_read_excel, pandas_preview, pandas_to_table,
    pandas_eval, pandas_query, pandas_merge, pandas_groupby_agg,
    pandas_fillna, pandas_rename, pandas_write_excel, pandas_ensure_columns, pandas_auto_rename, pandas_populate_vr,
    ])
    return {"llm": bound}


def run_vr_agent(
    base_path: str = DEFAULT_INPUT_PATH,
    template_path: str = DEFAULT_TEMPLATE_PATH,
    output_path: str = DEFAULT_OUTPUT_PATH,
    config: Optional[RunnableConfig] = None,
    competencia: Optional[str] = None,
    vigencia_start: Optional[str] = None,
    desligamento_cutoff: Optional[str] = None,
):
    logger = _get_logger()
    logger.info(
        "Starting VR agent; base=%s template=%s output=%s comp=%s vigencia_start=%s cutoff=%s",
        base_path, template_path, output_path, competencia, vigencia_start, desligamento_cutoff,
    )
    tools = [read_excel_table, read_excel_headers, write_excel_table, audit_log]
    # Add pandas tools if available
    tools += [pandas_read_excel, pandas_preview, pandas_to_table,
              pandas_eval, pandas_query, pandas_merge, pandas_groupby_agg,
              pandas_fillna, pandas_rename, pandas_write_excel, pandas_ensure_columns, pandas_auto_rename, pandas_populate_vr]
    agent = build_agent(tools)
    llm = agent["llm"]
    logger.info("Agent initialized and tools bound: %s", ", ".join([t.name for t in tools]))

    tool_map = {t.name: t for t in tools}

    # Provide file paths in the config metadata so the model can reference them explicitly
    meta = {
        "base_path": base_path,
        "template_path": template_path,
        "output_path": output_path,
        "competencia": competencia,
        "vigencia_start": vigencia_start,
        "desligamento_cutoff": desligamento_cutoff,
    }
    cfg = config or {}
    if isinstance(cfg, dict):
        cfg = {**cfg, "metadata": {**cfg.get("metadata", {}), **meta}}

    # Kick off with an initial instruction describing file paths
    # Build month context for the system prompt
    month_context = ""
    if competencia:
        month_context += f"- Competência: {competencia}.\n"
    if vigencia_start:
        month_context += f"- Vigência inicia em {vigencia_start}.\n"
    if desligamento_cutoff:
        month_context += f"- Regra de desligamento: zerar VR se data de desligamento <= {desligamento_cutoff}.\n"
    if not month_context:
        month_context = "- Vigência starts on 15/04.\n- Apply desligamento rules for May (05/2025).\n"

    user_msg = (
        "Load the base with pandas_read_excel from '"
        + os.path.relpath(base_path, BASE_DIR)
    + "'. Then call pandas_auto_rename followed by pandas_populate_vr to compute and fill all required columns. "
    + (f"Use pandas_populate_vr(competencia='{competencia}', vigencia_start='{vigencia_start}') " if competencia or vigencia_start else "")
    + "After that, enforce the exact layout with pandas_ensure_columns and save to '"
        + os.path.relpath(output_path, BASE_DIR)
        + "' using pandas_write_excel (preferred) or write_excel_table."
    )

    messages: List = [
        SystemMessage(content=SYSTEM_PROMPT_BASE.format(month_context=month_context)),
        HumanMessage(content=user_msg),
    ]

    final_text: Optional[str] = None
    wrote_output: bool = False
    output_path_returned: Optional[str] = None
    for i in range(20):  # safety cap on iterations
        t0 = _time.time()
        logger.info("Iteration %d: invoking LLM", i + 1)
        ai: AIMessage = llm.invoke(messages, config=cfg)
        dt = _time.time() - t0
        logger.info("Iteration %d completed in %.2fs", i + 1, dt)
        messages.append(ai)
        # If the model requested tools, execute them
        if getattr(ai, "tool_calls", None):
            logger.info("LLM requested %d tool call(s)", len(ai.tool_calls))
            for tc in ai.tool_calls:
                name = tc.get("name")
                args = tc.get("args", {})
                try:
                    arg_keys = list(args.keys()) if isinstance(args, dict) else []
                except Exception:
                    arg_keys = []
                logger.info("Tool call -> %s(args_keys=%s)", name, arg_keys)
                tool = tool_map.get(name)
                if not tool:
                    tool_result = f"ERROR: Unknown tool '{name}'."
                else:
                    try:
                        tool_result = tool.invoke(args)
                        preview = str(tool_result)
                        if len(preview) > 300:
                            preview = preview[:300] + "..."
                        logger.info("Tool result <- %s: %s", name, preview.replace("\n", " "))
                        if name in ("write_excel_table", "pandas_write_excel") and isinstance(tool_result, str) and not tool_result.startswith("ERROR"):
                            wrote_output = True
                            output_path_returned = tool_result
                    except Exception as e:
                        tool_result = f"ERROR: {e}"
                        logger.exception("Tool %s raised an exception", name)
                messages.append(ToolMessage(content=str(tool_result), tool_call_id=tc.get("id", "")))
            if wrote_output:
                final_text = f"OUTPUT_SAVED:{output_path_returned}"
                logger.info("Output file written successfully: %s", output_path_returned)
                break
            continue
        # No tool calls -> assume final response
        final_text = ai.content if isinstance(ai.content, str) else str(ai.content)
        content_len = len(str(final_text))
        logger.info("LLM returned final response (length=%d)", content_len)
        # If model returned empty content and no output was written, prompt it to continue properly
        if (not final_text or not str(final_text).strip()) and not wrote_output:
            logger.warning("Empty final response received; prompting model to continue and write output file.")
            messages.append(HumanMessage(content=(
                "Your last response was empty. Continue the task. "
                "You must save the file by calling either pandas_write_excel (preferred) or write_excel_table with valid headers_json and rows_json."
            )))
            continue
        break

    if not wrote_output:
        logger.error("Agent finished without writing the output file. Check logs and audit trail for details.")
    if final_text is None:
        logger.warning("Agent finished without a final textual response.")
    return final_text or ("OUTPUT_SAVED:" + output_path_returned if output_path_returned else "")


if __name__ == "__main__":
    try:
        result = run_vr_agent()
        print(result)
    except Exception as e:
        print(f"[ERROR] {e}")
